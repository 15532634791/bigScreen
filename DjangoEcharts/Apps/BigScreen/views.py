import os
import threading
from pathlib import Path
from openpyxl import load_workbook

from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet
from django.db.models import Q, Count, Avg, F, Subquery, OuterRef

from .serializers import BigScreenSerializer
from .serializers import FileUploadSerializer
from .models import BigScreen
from .models import UploadFile
from .models import InstallMaintainScore
from .models import InternetQualityScore


class ParseExcelWriteToDB(object):
    """
    此处解析excel
    """

    # 项目路径
    absolute_path = Path(__file__).resolve().parent.parent.parent

    def get_uploaded_excel_name(self):
        """
        返回上传文件名绝对路径
        :return:
        """
        last_file = UploadFile.objects.last()
        filename = last_file.file.name
        absolute_filename = os.path.join(self.absolute_path, filename)
        # 获取最近上传的文件名
        return absolute_filename

    def update_screen_comprehensive_score(self, userid_list):

        for userid in userid_list:

            install_score = InstallMaintainScore.objects.filter(userid=userid).values('score')
            internet_score = InternetQualityScore.objects.filter(userid=userid).values('score')
            install_score = int(install_score[0].get('score')) if install_score else None
            internet_score = int(internet_score[0].get('score')) if internet_score else None
            # print(userid, install_score, internet_score)
            # 两个都有值
            if internet_score and install_score:
                comprehensive_score = (internet_score + install_score) / 2
            # 一个有值
            elif internet_score and not install_score:
                comprehensive_score = internet_score
            # 一个有值
            elif install_score and not internet_score:
                comprehensive_score = install_score
            # 没有值
            else:
                comprehensive_score = None
            # print(userid, install_score, internet_score, comprehensive_score)
            if comprehensive_score:
                BigScreen.objects.filter(userid=userid).update(
                    comprehensive_score=comprehensive_score,
                    scoring_by_install_and_maintain_services=install_score,
                    internet_quality_rating=internet_score
                )

    @staticmethod
    def convert_int(value):
        try:
            return int(value)
        except Exception as e:
            return None

    def write_to_big_screen(self, workbook, worksheet):
        # 获取特定的表单
        ws = workbook[worksheet]  # 或者 wb['Sheet1'] 如果你知道表单名字
        # 遍历表单的行
        data_list = []
        for number, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # 处理每一行
            print(row)
            if number != 1:
                internet_quality_rating = self.parse_value(row[7])
                scoring_by_install_and_maintain_services = self.parse_value(row[8])
                # 两个都有值
                if internet_quality_rating and scoring_by_install_and_maintain_services:
                    comprehensive_score = (internet_quality_rating + scoring_by_install_and_maintain_services) / 2
                # 一个有值
                elif internet_quality_rating and not scoring_by_install_and_maintain_services:
                    comprehensive_score = internet_quality_rating
                # 一个有
                elif scoring_by_install_and_maintain_services and not internet_quality_rating:
                    comprehensive_score = scoring_by_install_and_maintain_services
                # 没有值
                else:
                    comprehensive_score = None

                big_screen = BigScreen(
                    county=row[0],
                    grid=row[1],
                    residential_quarters=row[2],
                    zone_of_responsibility=row[3],
                    packaging_and_maintenance=row[4],
                    userid=row[5],
                    userid_status=row[6],
                    internet_quality_rating=internet_quality_rating,
                    scoring_by_install_and_maintain_services=scoring_by_install_and_maintain_services,
                    comprehensive_score=comprehensive_score
                )
                data_list.append(big_screen)

        BigScreen.objects.bulk_create(data_list)

    def handle_screen_data(self, filename: str):
        """
        处理大屏数据
        """
        wb = load_workbook(filename=filename, read_only=True)
        try:
            sheet_names = wb.sheetnames
            data_list = []
            userid_list = []
            for worksheet in sheet_names:
                ws = wb[worksheet]
                for number, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if (number != 1) and row and (len(row) > 1):
                        print(number, row)
                        internet_score = self.convert_int(row[7])
                        install_score = self.convert_int(row[8])
                        userid = row[5]
                        userid_list.append(userid)
                        # 两个都有值
                        if internet_score and install_score:
                            comprehensive_score = (internet_score + install_score) / 2
                        # 一个有值
                        elif internet_score and not install_score:
                            comprehensive_score = internet_score
                        # 一个有值
                        elif install_score and not internet_score:
                            comprehensive_score = install_score
                        # 没有值
                        else:
                            comprehensive_score = None

                        big_screen = BigScreen(
                            county=row[0],
                            grid=row[1],
                            residential_quarters=row[2],
                            zone_of_responsibility=row[3],
                            packaging_and_maintenance=row[4],
                            userid=userid,
                            userid_status=row[6],
                            internet_quality_rating=internet_score,
                            scoring_by_install_and_maintain_services=install_score,
                            comprehensive_score=comprehensive_score
                        )
                        data_list.append(big_screen)

            BigScreen.objects.bulk_create(data_list)
            distinct_userid_list = set(userid_list)
            self.update_screen_comprehensive_score(distinct_userid_list)

        except Exception as e:
            pass
        finally:
            wb.close()

    def handle_score_data(self, filename: str):
        """
        处理评分数据
        """
        # 打开工作簿，使用只读模式
        wb = load_workbook(filename=filename, read_only=True)
        sheet_names = wb.sheetnames
        userid_list = []
        for worksheet in sheet_names:
            ws = wb[worksheet]
            if worksheet == '装维服务打分':
                for number, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if (number != 1) and row and (len(row) > 1):
                        userid = row[0]
                        score = row[1]
                        print('装维服务打分', number, userid, score)
                        userid_list.append(userid)
                        # 如果有多个 则更新最新的值
                        InstallMaintainScore.objects.update_or_create(
                            userid=userid,
                            defaults={'score': score}
                        )

            if worksheet == '上网质量打分':
                for number, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if (number != 1) and row and (len(row) > 1):
                        userid = row[0]
                        score = row[1]
                        print('上网质量打分', number, userid, score)
                        userid_list.append(userid)
                        InternetQualityScore.objects.update_or_create(
                            userid=userid,
                            defaults={'score': score}
                        )
        distinct_userid_list = set(userid_list)
        self.update_screen_comprehensive_score(distinct_userid_list)
        wb.close()

    def parse_excel(self):
        filename = self.get_uploaded_excel_name()

        if "打分数据" in filename:
            self.handle_score_data(filename=filename)

        elif "大屏数据" in filename:
            self.handle_screen_data(filename=filename)
        else:
            print({"status": False})

            return Response({"status": False})

        # self.update_screen_model()
        print("写入数据库完成")


class UploadFileViewSet(viewsets.ModelViewSet, ParseExcelWriteToDB):
    permission_classes = []
    queryset = UploadFile.objects.all()
    serializer_class = FileUploadSerializer

    def perform_create(self, serializer):
        # 调用父类的 perform_create 方法来保存对象
        super().perform_create(serializer)
        # 打印一条消息表示文件已成功上传
        print("文件上传成功")
        thread = threading.Thread(target=self.parse_excel)
        thread.start()


class BigScreenView(ModelViewSet):
    permission_classes = []
    queryset = BigScreen.objects.all()
    serializer_class = BigScreenSerializer


class InternetQualitySatisfaction(APIView):
    """
    上网质量满意度
    """
    permission_classes = []

    def get(self, request):
        # 上网质量打分：十分的个数
        ten_points_count = BigScreen.objects.filter(
            internet_quality_rating__in=[10]
        ).count()

        # 上网质量打分：七到九分的个数
        seven_to_nine_points_count = BigScreen.objects.filter(
            internet_quality_rating__in=[7, 8, 9]
        ).count()

        # 上网质量打分：一到六分的个数
        one_to_six_points_count = BigScreen.objects.filter(
            internet_quality_rating__in=[1, 2, 3, 4, 5, 6]
        ).count()

        # 上网质量打分：其他的个数
        other_count = BigScreen.objects.filter(
            Q(internet_quality_rating__isnull=True) | Q(internet_quality_rating="0分")
        ).count()

        data = {
            'ten_points_count': ten_points_count,
            'seven_to_nine_points_count': seven_to_nine_points_count,
            'one_to_six_points_count': one_to_six_points_count,
            'other_count': other_count
        }

        return Response(data)


class InstallAndMaintainPersonnelSatisfaction(APIView):
    """
    装维人员满意度
    """

    permission_classes = []

    def get(self, request):
        # 装维人员打分: 总个数

        # 装维人员打分：十分的个数
        ten_points_count = BigScreen.objects.filter(
            scoring_by_install_and_maintain_services__in=[10]
        ).count()

        # 装维人员打分：七到九分的个数
        seven_to_nine_points_count = BigScreen.objects.filter(
            scoring_by_install_and_maintain_services__in=[7, 8, 9]
        ).count()

        # 装维人员打分：一到六分的个数
        one_to_six_points_count = BigScreen.objects.filter(
            scoring_by_install_and_maintain_services__in=[1, 2, 3, 4, 5, 6]
        ).count()

        # 装维人员打分：其他的个数
        other_count = BigScreen.objects.filter(
            Q(scoring_by_install_and_maintain_services__isnull=True) | Q(scoring_by_install_and_maintain_services__in=[0])
        ).count()

        data = {
            'ten_points_count': ten_points_count,
            'seven_to_nine_points_count': seven_to_nine_points_count,
            'one_to_six_points_count': one_to_six_points_count,
            'other_count': other_count
        }
        return Response(data)


class InternetQualityProportion(APIView):
    """
    上网质量占比
    """
    permission_classes = []

    def get(self, request):
        # 获取所有区县
        try:
            county_list = list(
                BigScreen.objects.filter(county__isnull=False).values_list('county', flat=True).distinct())
        except Exception as e:
            county_list = []

        # 使用annotate和aggregate进行聚合查询，提高效率
        data = BigScreen.objects.filter(county__in=county_list).values('county') \
            .annotate(
            ten_points_count=Count('id', filter=Q(internet_quality_rating=10)),
            seven_to_nine_points_count=Count('id', filter=Q(internet_quality_rating__in=[7, 8, 9])),
            one_to_six_points_count=Count('id', filter=Q(internet_quality_rating__in=[1, 2, 3, 4, 5, 6])),
            other_count=Count('id', filter=Q(internet_quality_rating__isnull=True) | Q(internet_quality_rating__in=[0]))
        )

        county_list = [item['county'] for item in data]
        one_to_six_points_count_list = [item['one_to_six_points_count'] for item in data]
        seven_to_nine_points_count_list = [item['seven_to_nine_points_count'] for item in data]
        ten_points_count_list = [item['ten_points_count'] for item in data]
        other_count_list = [item['other_count'] for item in data]

        result = {
            'county_list': county_list[:15],
            'one_to_six_points_count_list': one_to_six_points_count_list[:15],
            'seven_to_nine_points_count_list': seven_to_nine_points_count_list[:15],
            'ten_points_count_list': ten_points_count_list[:15],
            'other_count_list': other_count_list[:15]
        }

        return Response(result)


class InstallAndMaintainProportion(APIView):
    """
    装维服务占比
    """
    permission_classes = []

    def get(self, request):
        # 获取所有区县
        try:
            county_list = list(
                BigScreen.objects.filter(county__isnull=False).values_list('county', flat=True).distinct())
        except Exception as e:
            county_list = []

        # 使用annotate和aggregate进行聚合查询，提高效率
        data = BigScreen.objects.filter(county__in=county_list).values('county') \
            .annotate(
            ten_points_count=Count('id', filter=Q(scoring_by_install_and_maintain_services__in=[10])),
            seven_to_nine_points_count=Count('id', filter=Q(
                scoring_by_install_and_maintain_services__in=[7, 8, 9])),
            one_to_six_points_count=Count('id', filter=Q(
                scoring_by_install_and_maintain_services__in=[1, 2, 3, 4, 5, 6])),
            other_count=Count('id', filter=Q(scoring_by_install_and_maintain_services__isnull=True) | Q(
                scoring_by_install_and_maintain_services__in=[0]))
        )

        data = {
            'county_list': [item['county'] for item in data][:15],
            'ten_points_count': [item['ten_points_count'] for item in data][:15],
            "seven_to_nine_points_count": [item['seven_to_nine_points_count'] for item in data][:15],
            'one_to_six_points_count': [item['one_to_six_points_count'] for item in data][:15],
            'other_count': [item['other_count'] for item in data]

        }

        return Response(data)


class CountySatisfaction(APIView):
    """
    区县满意度
    """
    permission_classes = []

    def get(self, request):

        # 获取所有区县
        try:
            counties = list(BigScreen.objects.filter(county__isnull=False).values_list('county', flat=True).distinct())
        except Exception as e:
            counties = []

        comprehensive_score_list = []

        for county in counties:
            county_count = BigScreen.objects.filter(
                # 匹配区县不为空的，所有值的平均值
                county=county, comprehensive_score__isnull=False
            ).aggregate(
                count=Count('id'),
                average_comprehensive_score=Avg('comprehensive_score')
            )

            count = county_count['count']
            average_score = county_count['average_comprehensive_score']

            # print(count, average_score)
            try:
                average_score = (average_score - 1) / 9 * 100
            except TypeError:
                average_score = 0
            county_average_score = {
                "county": county, "average_score": average_score
            }
            if average_score > 0:
                comprehensive_score_list.append(county_average_score)
        sorted_data = sorted(comprehensive_score_list, key=lambda x: x['average_score'], reverse=True)
        top_15 = sorted_data[:15]

        top_15_counties = []
        top_15_score = []
        for top in top_15:
            top_15_counties.append(top.get('county'))
            top_15_score.append(top.get('average_score'))

        data = {
            "comprehensive_score_list": top_15_score,
            "counties": top_15_counties
        }

        return Response(data)


class GridSatisfaction(APIView):
    """
    网格满意度
    """
    permission_classes = []

    def get(self, request):
        # 获取所有区县
        try:
            grids = list(BigScreen.objects.filter(grid__isnull=False).values_list('grid', flat=True).distinct())
        except Exception as e:
            grids = []

        comprehensive_score_list = []

        for grid in grids:
            grid_count = BigScreen.objects.filter(
                grid=grid, comprehensive_score__isnull=False
            ).aggregate(
                count=Count('id'),
                average_comprehensive_score=Avg('comprehensive_score')
            )
            average_score = grid_count['average_comprehensive_score']
            try:
                average_score = (average_score - 1) / 9 * 100
            except TypeError:
                average_score = 0
            print(grid)
            grid_average_score = {
                "grid": grid, "average_score": average_score
            }
            if average_score > 0:
                comprehensive_score_list.append(grid_average_score)

        sorted_data = sorted(comprehensive_score_list, key=lambda x: x['average_score'])
        top_15 = sorted_data[:15]

        top_15_grids = []
        top_15_score = []
        for top in top_15:
            top_15_grids.append(top.get('grid'))
            top_15_score.append(top.get('average_score'))

        data = {
            "comprehensive_score_list": top_15_score,
            "grids": top_15_grids
        }

        return Response(data)


class ResponsibilityZoneSatisfaction(APIView):
    """
    责任区
    """
    permission_classes = []

    def get(self, request):
        try:
            zones = list(
                BigScreen.objects.filter(zone_of_responsibility__isnull=False).values_list('zone_of_responsibility',
                                                                                           flat=True).distinct()
            )
        except Exception as e:
            zones = []

        comprehensive_score_list = []

        for zone in zones:
            zone_count = BigScreen.objects.filter(
                zone_of_responsibility=zone, comprehensive_score__isnull=False
            ).aggregate(
                count=Count('id'),
                average_comprehensive_score=Avg('comprehensive_score')
            )
            average_score = zone_count['average_comprehensive_score']
            try:
                average_score = (average_score - 1) / 9 * 100
            except TypeError:
                average_score = 0
            zone_average_score = {
                "zone": zone, "average_score": average_score
            }
            if average_score > 0:
                comprehensive_score_list.append(zone_average_score)

        sorted_data = sorted(comprehensive_score_list, key=lambda x: x['average_score'])
        top_15 = sorted_data[:15]

        top_15_zones = []
        top_15_score = []
        for top in top_15:
            top_15_zones.append(top.get('zone'))
            top_15_score.append(top.get('average_score'))

        data = {
            "comprehensive_score_list": top_15_score,
            "zones": top_15_zones
        }

        # data = {
        #     "comprehensive_score_list": comprehensive_score_list[:15],
        #     "zones": zones[:15]
        # }

        return Response(data)


class GridRank(APIView):
    permission_classes = []

    def get(self, request):
        # 查找网格
        search = request.query_params.get("value")
        # # 获取所有网格
        try:
            grids = list(BigScreen.objects.filter(grid__isnull=False).values_list('grid', flat=True).distinct())
        except Exception as e:
            print(e)
            grids = []

        label = []
        for grid in grids:
            label.append({"value": grid, "label": grid})

        comprehensive_score_list = []

        score = None

        for grid in grids:
            grid_count = BigScreen.objects.filter(
                grid=grid, comprehensive_score__isnull=False
            ).aggregate(
                count=Count('id'),
                average_comprehensive_score=Avg('comprehensive_score')
            )
            average_score = grid_count['average_comprehensive_score']
            try:
                average_score = (average_score - 1) / 9 * 100
            except TypeError:
                average_score = 0
            if grid == search:
                score = round(average_score, 2)
            comprehensive_score_list.append(
                {grid: average_score}
            )

        comprehensive_score_list.sort(key=lambda x: list(x.values())[0], reverse=True)

        rank = None
        for index, d in enumerate(comprehensive_score_list):
            if search in d.keys():
                # print(f"The index of {search} is: {index}")
                rank = index + 1
                break

        internet_count = BigScreen.objects.filter(grid=search).values('grid') \
            .annotate(
            ten_points_count=Count('id', filter=Q(internet_quality_rating__in=[10])),
            seven_to_nine_points_count=Count('id', filter=Q(internet_quality_rating__in=[7, 8, 9])),
            one_to_six_points_count=Count('id', filter=Q(internet_quality_rating__in=[1, 2, 3, 4, 5, 6])),
            other_count=Count('id', filter=Q(internet_quality_rating__isnull=True))
        )
        # print(internet_count)
        internet = {
            'ten_points_count': [item['ten_points_count'] for item in internet_count],
            "seven_to_nine_points_count": [item['seven_to_nine_points_count'] for item in internet_count],
            'one_to_six_points_count': [item['one_to_six_points_count'] for item in internet_count],
            'other_count': [item['other_count'] for item in internet_count]
        }

        install_count = BigScreen.objects.filter(grid=search).values('grid') \
            .annotate(
            ten_points_count=Count('id', filter=Q(scoring_by_install_and_maintain_services__in=[10])),
            seven_to_nine_points_count=Count('id', filter=Q(scoring_by_install_and_maintain_services__in=[7, 8, 9])),
            one_to_six_points_count=Count('id',
                                          filter=Q(scoring_by_install_and_maintain_services__in=[1, 2, 3, 4, 5, 6])),
            other_count=Count('id', filter=Q(
                scoring_by_install_and_maintain_services__isnull=True))
        )
        install = {
            'ten_points_count': [item['ten_points_count'] for item in install_count],
            "seven_to_nine_points_count": [item['seven_to_nine_points_count'] for item in install_count],
            'one_to_six_points_count': [item['one_to_six_points_count'] for item in install_count],
            'other_count': [item['other_count'] for item in install_count]
        }

        data = {
            'rank': rank, 'label': label, "internet": internet, "install": install, "score": score
        }
        return Response(data)


class CountyRank(APIView):
    permission_classes = []

    def get(self, request):
        # 查找网格
        search = request.query_params.get("value")
        # # 获取所有网格
        try:
            counties = list(BigScreen.objects.filter(county__isnull=False).values_list('county', flat=True).distinct())
        except Exception as e:
            print(e)
            counties = []

        label = []

        for county in counties:
            label.append({"value": county, "label": county})

        comprehensive_score_list = []

        score = None
        for county in counties:
            county_count = BigScreen.objects.filter(
                county=county, comprehensive_score__isnull=False
            ).aggregate(
                count=Count('id'),
                average_comprehensive_score=Avg('comprehensive_score')
            )
            average_score = county_count['average_comprehensive_score']
            try:
                average_score = (average_score - 1) / 9 * 100
            except TypeError:
                average_score = 0
            if county == search:
                score = round(average_score, 2)
            comprehensive_score_list.append(
                {county: average_score}
            )

        comprehensive_score_list.sort(key=lambda x: list(x.values())[0], reverse=True)

        rank = None
        for index, d in enumerate(comprehensive_score_list):
            if search in d.keys():
                # print(f"The index of {search} is: {index}")
                rank = index + 1
                break

        internet_count = BigScreen.objects.filter(county=search).values('county') \
            .annotate(
            ten_points_count=Count('id', filter=Q(internet_quality_rating__in=[10])),
            seven_to_nine_points_count=Count('id', filter=Q(internet_quality_rating__in=[7, 8, 9])),
            one_to_six_points_count=Count('id', filter=Q(internet_quality_rating__in=[1, 2, 3, 4, 5, 6])),
            other_count=Count('id', filter=Q(internet_quality_rating__isnull=True))
        )

        internet = {
            'ten_points_count': [item['ten_points_count'] for item in internet_count],
            "seven_to_nine_points_count": [item['seven_to_nine_points_count'] for item in internet_count],
            'one_to_six_points_count': [item['one_to_six_points_count'] for item in internet_count],
            'other_count': [item['other_count'] for item in internet_count]
        }

        install_count = BigScreen.objects.filter(county=search).values('county') \
            .annotate(
            ten_points_count=Count('id', filter=Q(scoring_by_install_and_maintain_services__in=[10])),
            seven_to_nine_points_count=Count('id', filter=Q(scoring_by_install_and_maintain_services__in=[7, 8, 9])),
            one_to_six_points_count=Count('id',
                                          filter=Q(scoring_by_install_and_maintain_services__in=[1, 2, 3, 4, 5, 6])),
            other_count=Count('id', filter=Q(
                scoring_by_install_and_maintain_services__isnull=True))
        )
        install = {
            'ten_points_count': [item['ten_points_count'] for item in install_count],
            "seven_to_nine_points_count": [item['seven_to_nine_points_count'] for item in install_count],
            'one_to_six_points_count': [item['one_to_six_points_count'] for item in install_count],
            'other_count': [item['other_count'] for item in install_count]
        }

        data = {
            'rank': rank, 'label': label, "internet": internet, "install": install, "score": score
        }

        return Response(data)


class ResponsibilityZoneRank(APIView):
    permission_classes = []

    def get(self, request):
        # 查找网格
        search = request.query_params.get("value")
        # # 获取所有网格
        try:
            zones = list(
                BigScreen.objects.filter(zone_of_responsibility__isnull=False).values_list('zone_of_responsibility',
                                                                                           flat=True).distinct())
        except Exception as e:
            print(e)
            zones = []
        # print(zones)
        label = []

        for zone in zones:
            label.append({"value": zone, "label": zone})

        comprehensive_score_list = []
        score = None

        for zone in zones:
            county_count = BigScreen.objects.filter(
                zone_of_responsibility=zone, comprehensive_score__isnull=False
            ).aggregate(
                count=Count('id'),
                average_comprehensive_score=Avg('comprehensive_score')
            )
            average_score = county_count['average_comprehensive_score']
            try:
                average_score = (average_score - 1) / 9 * 100
            except TypeError:
                average_score = 0
            if zone == search:
                score = round(average_score, 2)
            comprehensive_score_list.append(
                {zone: average_score}
            )

        comprehensive_score_list.sort(key=lambda x: list(x.values())[0], reverse=True)
        # print(comprehensive_score_list)

        rank = None
        for index, d in enumerate(comprehensive_score_list):
            if search in d.keys():
                # print(f"The index of {search} is: {index}")
                rank = index + 1
                break

        internet_count = BigScreen.objects.filter(zone_of_responsibility=search).values('zone_of_responsibility') \
            .annotate(
            ten_points_count=Count('id', filter=Q(internet_quality_rating__in=[10])),
            seven_to_nine_points_count=Count('id', filter=Q(internet_quality_rating__in=[7, 8, 9])),
            one_to_six_points_count=Count('id', filter=Q(internet_quality_rating__in=[1, 2, 3, 4, 5, 6])),
            other_count=Count('id', filter=Q(internet_quality_rating__isnull=True))
        )

        internet = {
            'ten_points_count': [item['ten_points_count'] for item in internet_count],
            "seven_to_nine_points_count": [item['seven_to_nine_points_count'] for item in internet_count],
            'one_to_six_points_count': [item['one_to_six_points_count'] for item in internet_count],
            'other_count': [item['other_count'] for item in internet_count]
        }

        install_count = BigScreen.objects.filter(zone_of_responsibility=search).values('zone_of_responsibility') \
            .annotate(
            ten_points_count=Count('id', filter=Q(scoring_by_install_and_maintain_services__in=[10])),
            seven_to_nine_points_count=Count('id', filter=Q(scoring_by_install_and_maintain_services__in=[7, 8, 9])),
            one_to_six_points_count=Count('id',
                                          filter=Q(scoring_by_install_and_maintain_services__in=[1, 2, 3, 4, 5, 6])),
            other_count=Count('id', filter=Q(
                scoring_by_install_and_maintain_services__isnull=True))
        )
        install = {
            'ten_points_count': [item['ten_points_count'] for item in install_count],
            "seven_to_nine_points_count": [item['seven_to_nine_points_count'] for item in install_count],
            'one_to_six_points_count': [item['one_to_six_points_count'] for item in install_count],
            'other_count': [item['other_count'] for item in install_count]
        }

        data = {
            'rank': rank, 'score': score, 'label': label, "internet": internet, "install": install
        }

        return Response(data)
